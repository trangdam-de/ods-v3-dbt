# Standard library imports
import datetime
import logging
from datetime import datetime, timedelta
import zipfile
from io import BytesIO

# Third-party library imports
import openpyxl
import psycopg
import pandas as pd
from airflow.hooks.base import BaseHook
from airflow.providers.amazon.aws.hooks.s3 import S3Hook
from airflow.models import BaseOperator
from airflow.utils.decorators import apply_defaults
from psycopg import sql
from soda.scan import Scan

from helper1.task_logger import TaskRunMetadata

LOGGER = logging.getLogger(__name__)

class PNSToStagingDailyOperator(BaseOperator):
    @apply_defaults
    def __init__(self,
                 staging_conn_id: str,
                 minio_conn_id: str,
                 minio_bucket_name: str,
                 des_schema_name: str,
                 des_table_name: str,
                 minio_key: str,
                 columns: dict,
                 file_prefix: str,
                 header_rows: int,
                 end_skip: int,
                 type: str,
                 cursor_field: str = None,
                 batch_size: int = 50000,
                 *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.stg_connection = None
        self.src_file_path = None
        self.task_metadata = None
        self.minio_conn_id = minio_conn_id
        self.minio_bucket_name = minio_bucket_name
        self.staging_conn_id = staging_conn_id
        self.cursor_field = cursor_field
        self.des_schema_name = des_schema_name
        self.des_table_name = des_table_name
        self.columns = list(columns.keys())
        self.column_types = columns
        self.file_prefix = file_prefix
        self.header_rows = header_rows
        self.minio_key = minio_key
        self.file_dates = []
        self.end_skip = end_skip
        self.type = type
        self.batch_size = batch_size
        self.status_date = None

    def pre_execute(self, context):
        LOGGER.info(f"{self.task_id} - pre_execute() - START establishing source and staging connection")
        base_date = context['logical_date'].in_tz('Asia/Ho_Chi_Minh').start_of('day')
        if context["dag_run"].run_type == "scheduled":
            previous_date = base_date.strftime('%d%m%Y')
            self.file_dates.append(previous_date)
            self.status_date = base_date.strftime('%Y%m%d')
            #self.file_dates.append(base_date.strftime('%d%m%Y')
        else:
            self.start_time = datetime.strptime(context["dag_run"].conf.get("start_time"), "%Y-%m-%d %H:%M:%S")
            self.end_time = datetime.strptime(context["dag_run"].conf.get("end_time"), "%Y-%m-%d %H:%M:%S")
            current_date = self.start_time
            self.status_date = self.start_time.strftime('%Y%m%d')
            while current_date < self.end_time:
                file_date = current_date.strftime("%d%m%Y")
                self.file_dates.append(file_date)
                current_date += timedelta(days=1)
            if not self.start_time or not self.end_time:
                LOGGER.error(f"{self.task_id} - pre_execute() - |ERROR| --> there is no configuration for start_time and end_time")
                raise

        LOGGER.info(f"Type of status_date is: {type(self.status_date)}")
        context['ti'].xcom_push(key = "status_date", value = self.status_date)
        self.task_metadata = TaskRunMetadata(context=context,
                                             table_name=self.des_schema_name,
                                             result_connection=BaseHook.get_connection(
                                                 "staging_conn_id").get_uri())
        try:
            self.stg_connection = psycopg.connect(BaseHook.get_connection(self.staging_conn_id).get_uri())
        except Exception as e:
            LOGGER.error(f"{self.task_id} - pre_execute() - |ERROR| --> {e}", exc_info=True)
            raise
        LOGGER.info(f"{self.task_id} - pre_execute() - END connections established")

    def execute(self, context):
        LOGGER.info(f"{self.task_id} - execute() - START fetching and processing data")
        print(f"Processing {len(self.file_dates)} file dates")
        for self.file_date in self.file_dates:
            try:
                # Get file from MinIO
                selected_file, data = self.get_file_from_minio()
                
                # Load Excel file in read-only mode
                LOGGER.info(f"{self.task_id} - execute() - Attempting to load Excel file: {selected_file}")
                try:
                    # Ensure stream is at the beginning
                    if hasattr(data, 'seek'):
                        data.seek(0)
                    wb = openpyxl.load_workbook(data, read_only=True, data_only=True)
                except zipfile.BadZipFile as e:
                    LOGGER.error(f"{self.task_id} - execute() - |ERROR|: File {selected_file} is not a valid Excel file: {str(e)}")
                    self.task_metadata.write_result(result=f"Invalid Excel file: {selected_file}", is_success=False)
                    raise
                except Exception as e:
                    LOGGER.error(f"{self.task_id} - execute() - |ERROR|: Failed to load Excel file {selected_file}: {str(e)}", exc_info=True)
                    self.task_metadata.write_result(result=f"Failed to load Excel file: {selected_file}", is_success=False)
                    raise
                
                # Check if workbook has at least one sheet
                if not wb.worksheets:
                    LOGGER.error(f"{self.task_id} - execute() - |ERROR|: Workbook {selected_file} has no sheets")
                    self.task_metadata.write_result(result=f"No sheets in workbook: {selected_file}", is_success=False)
                    raise ValueError(f"Workbook {selected_file} has no sheets")
                
                # Select the first sheet dynamically
                ws = wb.worksheets[0]
                LOGGER.info(f"{self.task_id} - execute() - Processing sheet: {ws.title}")
                
                # Initialize batch processing
                batch = []
                row_count = 0
                total_rows = 0
                
                # Skip header rows and iterate through data rows
                for row_idx, row in enumerate(ws.iter_rows(min_row=self.header_rows + 1, 
                                                        max_col=len(self.columns), 
                                                        values_only=True), start=self.header_rows + 1):
                    # Skip rows at the end if end_skip is specified
                    if self.end_skip > 0 and row_idx > ws.max_row - self.end_skip:
                        continue
                            
                    # Clean and process row
                    cleaned_row = []
                    for idx, val in enumerate(row[:len(self.columns)]):
                        col_name = self.columns[idx]
                        col_type = self.column_types.get(col_name, 'str').lower()  
                        if col_type == 'Int64':
                            if val is None or val == "" or pd.isna(val) or val == "NULL" or str(val).strip().lower() == "nan" or val == '':
                                cleaned_row.append(None)
                            else:
                                try:
                                    #int_val = int(val)
                                    int_val = int(str(val).strip())
                                    if not (-2147483648 <= int_val <= 2147483647):
                                        LOGGER.error(
                                            f"{self.task_id} - NumericValueOutOfRange: Value '{val}' in column '{col_name}' "
                                        )
                                        cleaned_row.append(None)
                                    else:
                                        cleaned_row.append(int_val)
                                except (ValueError, TypeError):
                                    LOGGER.warning(f"{self.task_id} - Integer value is invalid '{val}' in '{col_name}' ")
                                    cleaned_row.append(None)
                        else:
                            cleaned_row.append(None if pd.isna(val) or val == "NULL" else val)
                    
                    # Apply type-specific cleaning
                    if self.type == 'category' and self.des_table_name == 'collection_delivery_route':
                        if cleaned_row[self.columns.index('route_code')] is None or \
                        cleaned_row[self.columns.index('unit_code')] is None:
                            continue  # Skip rows with null route_code or unit_code
                    elif self.type == 'detail':
                        if cleaned_row[self.columns.index('lading_code')] is None:
                            continue  # Skip rows with null lading_code
                        cleaned_row.append(pd.to_datetime(datetime.today()))  # Add etl_date
                    
                    batch.append(cleaned_row)
                    row_count += 1
                    total_rows += 1
                    
                    # Process batch when it reaches batch_size
                    if len(batch) >= self.batch_size:
                        LOGGER.info(f"{self.task_id} - Processing batch of {len(batch)} rows (Total: {total_rows})")
                        
                        # Validate schema for the batch
                        df = pd.DataFrame(batch, columns=self.columns + (['etl_date'] if self.type == 'detail' else []))
                        if self.type == 'category' and self.des_table_name == 'collection_delivery_route':
                            df = df.drop_duplicates(subset=["route_code", "unit_code"], keep="first")
                        self.validate_schema(dataframe=df, context=context)
                        
                        # Load batch into PostgreSQL
                        self.load_data(dataframe=df, context=context)
                        
                        # Clear batch
                        batch = []
                        row_count = 0
                    
                    # Log progress every 100,000 rows
                    if total_rows % 100000 == 0:
                        LOGGER.info(f"{self.task_id} - Processed {total_rows} rows")
                
                # Process remaining rows in the last batch
                if batch:
                    LOGGER.info(f"{self.task_id} - Processing final batch of {len(batch)} rows (Total: {total_rows})")
                    df = pd.DataFrame(batch, columns=self.columns + (['etl_date'] if self.type == 'detail' else []))
                    if self.type == 'category' and self.des_table_name == 'collection_delivery_route':
                        df = df.drop_duplicates(subset=["route_code", "unit_code"], keep="first")
                    self.validate_schema(dataframe=df, context=context)
                    self.load_data(dataframe=df, context=context)
                
                LOGGER.info(f"{self.task_id} - Completed processing {selected_file} with {total_rows} rows")
                wb.close()
            
            except (zipfile.BadZipFile, ExcelReadError, FileNotFoundInMinIOException) as e:
                LOGGER.error(f"{self.task_id} - execute() - |ERROR|: Skipping file for date {self.file_date}: {str(e)}")
                continue  # Skip to the next file_date
            except Exception as e:
                LOGGER.error(f"{self.task_id} - execute() - |ERROR|: Failed processing file for date {self.file_date}: {str(e)}", exc_info=True)
                raise
        
        LOGGER.info(f"{self.task_id} - execute() - END loading data to staging layer")
        #return self.status_date

    def post_execute(self, context, result=None):
        LOGGER.info(f"{self.task_id} - post_execute() - START calling task callback")
        self.stg_connection.close()
        self.task_metadata.write_result(result=result, is_success=True)
        LOGGER.info(f"{self.task_id} - post_execute() - END calling task callback --> See the task results")

    def get_file_from_minio(self) -> tuple[str, BytesIO]:
        LOGGER.info(f"{self.task_id} - get_file_from_minio() - START getting xlsx file from MinIO")
        try:
            s3_hook = S3Hook(aws_conn_id=self.minio_conn_id)
            objects = s3_hook.list_keys(bucket_name=self.minio_bucket_name, prefix=self.minio_key)
            LOGGER.info(f"{self.task_id} - Found {len(objects)} objects in {self.minio_key}: {objects}")
            matching_files = []
            
            for obj_key in objects:
                file_name = obj_key.split('/')[-1]
                # Strict check for .xlsx extension
                if not file_name.lower().endswith('.xlsx'):
                    LOGGER.warning(f"{self.task_id} - Skipping non-xlsx file: {file_name}")
                    continue
                if (self.type == 'detail' and
                        self.file_prefix.lower() in file_name.lower() and
                        self.file_date in file_name.lower()):
                    matching_files.append(obj_key)
                elif (self.type == 'category' and
                      self.file_prefix.lower() in file_name.lower()):
                    matching_files.append(obj_key)
            
            if not matching_files:
                error_msg = (f"File xlsx không tồn tại với prefix '{self.file_prefix}' "
                             f"và ngày '{self.file_date}' trong folder {self.minio_key}.")
                LOGGER.error(f"{self.task_id} - |ERROR| {error_msg}")
                self.task_metadata.write_result(result="Xlsx file does not exist", is_success=False)
                raise FileNotFoundInMinIOException(error_msg)
            
            # Select the first matching file
            selected_file = matching_files[0]
            LOGGER.info(f"{self.task_id} - Selected file: {selected_file}")
            
            # Get file as StreamingBody
            s3_object = s3_hook.get_key(key=selected_file, bucket_name=self.minio_bucket_name)
            file_data = s3_object.get()["Body"]
            
            # Buffer the stream to validate without consuming
            file_content = file_data.read()
            if not file_content:
                LOGGER.error(f"{self.task_id} - |ERROR| File {selected_file} is empty")
                self.task_metadata.write_result(result=f"Empty file: {selected_file}", is_success=False)
                raise ExcelReadError(f"File {selected_file} is empty")
            
            # Log first few bytes for debugging
            LOGGER.debug(f"{self.task_id} - First 10 bytes of file {selected_file}: {file_content[:10]}")
            
            # Validate ZIP format
            file_io = BytesIO(file_content)
            try:
                with zipfile.ZipFile(file_io, 'r') as zf:
                    if zf.testzip() is not None:
                        LOGGER.error(f"{self.task_id} - |ERROR| File {selected_file} is a corrupted ZIP/Excel file")
                        self.task_metadata.write_result(result=f"Corrupted ZIP/Excel file: {selected_file}", is_success=False)
                        raise zipfile.BadZipFile(f"File {selected_file} is a corrupted ZIP/Excel file")
            except zipfile.BadZipFile as e:
                LOGGER.error(f"{self.task_id} - |ERROR| File {selected_file} is not a valid ZIP/Excel file: {str(e)}")
                self.task_metadata.write_result(result=f"Invalid ZIP/Excel file: {selected_file}", is_success=False)
                raise
            
            # Create a new BytesIO for openpyxl
            file_io = BytesIO(file_content)
            LOGGER.info(f"{self.task_id} - get_file_from_minio() - Successfully read file: {selected_file} (size: {len(file_content)} bytes)")
            return selected_file, file_io
            
        except Exception as e:
            LOGGER.error(f"{self.task_id} - |ERROR| Error while getting file from MinIO: {str(e)}", exc_info=True)
            self.task_metadata.write_result(result=f"Error while getting file from MinIO: {str(e)}", is_success=False)
            raise

    def validate_schema(self, dataframe, context):
        LOGGER.info(f"{self.task_id} - validate_schema() - START validating source data")
        scan = Scan()
        try:
            scan.set_scan_definition_name(self.des_schema_name)
            scan.set_data_source_name("pandas")
            scan.add_pandas_dataframe(dataset_name=self.des_table_name, pandas_df=dataframe, data_source_name="pandas")
            scan.add_sodacl_yaml_file(f"/opt/airflow/soda/check/{self.des_schema_name}/{self.des_table_name}.yml")
            scan.set_verbose(True)
            scan.execute()
            scan.assert_no_error_logs()
            scan.assert_no_checks_fail()
            LOGGER.info(f"{self.task_id} - validate_schema() - END validating source data")
        except Exception as e:
            self.task_metadata.write_result(result="Fail while validating data", is_success=False)
            LOGGER.error(f"{self.task_id} - validate_schema() - |ERROR| --> {str(e)}", exc_info=True)
            if scan.has_error_logs():
                LOGGER.error(f"{self.task_id} - validate_schema() - |ERROR| --> {scan.get_error_logs_text()}")
            if scan.has_check_fails():
                LOGGER.error(f"{self.task_id} - validate_schema() - |ERROR| --> {scan.get_checks_fail_text()}")
            raise

    def load_data(self, dataframe, context):
        LOGGER.info(f"{self.task_id} - load_data() - START loading data to staging layer")

        for col in ['transport_type', 'collection_main', 'contract_type', 'quantity', 'so_tien_thu_ho', 'service_code']:
            if col in dataframe.columns:
                dataframe[col] = dataframe[col].apply(lambda x: None if x == "" else x)

        try:
            columns = dataframe.columns.map(str.lower).tolist()
            insert_query = sql.SQL(
                "INSERT INTO {} ({}) VALUES ({})"
            ).format(
                sql.Identifier("staging", f"{self.des_schema_name}_{self.des_table_name}"),
                sql.SQL(", ").join(map(sql.Identifier, columns)),
                sql.SQL(", ").join(sql.Placeholder() * len(columns))
            )
            
            # Only truncate table for the first batch of the first file
            if not hasattr(self, '_table_truncated'):
                truncate_query = sql.SQL(
                    "TRUNCATE TABLE {};"
                ).format(
                    sql.Identifier("staging", f"{self.des_schema_name}_{self.des_table_name}")
                )
                with self.stg_connection.cursor() as cur:
                    cur.execute(truncate_query)
                self._table_truncated = True
            
            with self.stg_connection.cursor() as cur:
                cur.executemany(insert_query, dataframe.itertuples(index=False, name=None))
            self.stg_connection.commit()
            LOGGER.info(f"{self.task_id} - load_data() - END loading data to staging layer")
        except Exception as e:
            self.task_metadata.write_result(result="Fail while loadings data", is_success=False)
            LOGGER.error(f"{self.task_id} - load_data() - |ERROR| when loading data to staging layer --> {e}", exc_info=True)
            raise

class FileNotFoundErrorException(Exception):
    pass

class FileNotFoundInMinIOException(Exception):
    pass

class ExcelReadError(Exception):
    pass
